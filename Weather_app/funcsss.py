import requests
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from os import path
import config as c

def get_date_time(ts, timezone, dt_format="%H:%M:%S"):
    tz = datetime.timezone(datetime.timedelta(seconds=timezone))
    return datetime.datetime.fromtimestamp(ts, tz=tz).strftime(dt_format)


def get_weather(city_name):
    params = {
        "appid": c.API_KEY,
        "units": c.UNITS,
        "lang": c.LANG,
        "q": city_name
    }
    try:
        r = requests.get(c.API_URL, params=params)
        return r.json()
    except:
        return {"cod": 0, "message": "Failed data access"}


def print_weather(data):

    if data['cod'] != 200:
        print(data['message'])
        return {}
    else:
        sunrise_time = get_date_time(data['sys']['sunrise'], data['timezone'])
        sunset_time = get_date_time(data['sys']['sunset'], data['timezone'])
        print(f"""
        Localization: {data['name']}, Country: {data['sys']['country']}
        Temreture: {data['main']['temp']} ℃
        Pressure: {data['main']['pressure']} Pa
        Humidity: {data['main']['humidity']}%
        Speed of wind: {data['wind']['speed']} m/sec
        Weather conditions: {data['weather'][0]['description']}
        Sunrise: {sunrise_time}
        Sunset time: {sunset_time}
        """)

        print("+" * 50)
        return data


def save_excel(data):
    if data['cod'] == 200:
        if path.exists(c.FILE_EXCEL):
            wb = load_workbook(filename=c.FILE_EXCEL)
            ws = wb.active           
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Statistic of requests"
            ws.append(['Date of request', 'City', 'Temperature'])
            ft = Font(bold=True, color='0000FFFF')
            a1 = ws['A1']
            b1 = ws['B1']
            c1 = ws['C1']
            a1.font = b1.font = c1.font = ft

        ws.append([datetime.datetime.now(), f"{data['name']}, {data['sys']['country']}", data['main']['temp']])
        wb.save(filename=c.FILE_EXCEL)    
    else:
        pass